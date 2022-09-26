import sys
#from isbntools.app import *
import pandas
from isbnlib import meta
from isbnlib.registry import bibformatters
import openpyxl
import mysql.connector
from mysql.connector import Error

# load excel with its path
wrkbk = openpyxl.load_workbook("inventory.xlsx")
  
sh = wrkbk.active



# initialize lists
bad_list = []
good_list = []
genre_list = []
good_dict = {"isbn":[],"year":[],"publisher":[],"author":[],"title":[]}

try:
    connection = mysql.connector.connect(user='jfsharron', password='marie151414',
    host='192.168.2.107', database='isbn22')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = connection.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

# Input Variables
def createLists():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """

    # iterate through excel and display data
    for row in sh.iter_rows(min_row=2, min_col=1, max_row=63, max_col=1):
        for cell in row:
            isbn = str(cell.value)

            SERVICE = "openl"

            bibtex = bibformatters["bibtex"]
            
            meta_dict = meta(isbn, service='default')

            if meta_dict.get('Authors') is None:
                bad_list.append(isbn)
            else:
                good_list.append(isbn)

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=63, max_col=2):
        for cell in row:
            genre = str(cell.value)
            genre_list.append(genre)

    #print("bad list ", bad_list)
    #print("good list ", good_list)
    #print("genre list ", genre_list)

    # ==========================================================================

def getInfo():    
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    ============================================================================
    """
    for i in good_list:
        #for cell in row:
            isbn = i

            SERVICE = "openl"

            bibtex = bibformatters["bibtex"]
            
            meta_dict = meta(isbn, service='default')

            aut = str(meta_dict['Authors'])
            aut = aut.replace("[","")
            aut = aut.replace("]","")
            author = aut.replace("'","")
            title = meta_dict['Title']
            isbn = meta_dict['ISBN-13']
            year = meta_dict['Year']
            publisher = meta_dict['Publisher']


            data = (isbn, year, publisher, author, title)

            mySql_insert_query = (
            "INSERT INTO isbn (isbn, year, publisher, author, title)"
            "VALUES (%s, %s, %s, %s, %s)"
            )
        
            cursor.execute(mySql_insert_query, data)
            connection.commit()
            print(cursor.rowcount, "Record successfully inserted into isbn")








            #good_dict["isbn"].append(isbn)
            #good_dict["year"].append(year)
            #good_dict["publisher"].append(publisher)
            #good_dict["author"].append(author)
            #good_dict["title"].append(title)


    #for i in good_dict:
    #    mySql_insert_query = (
    #    "INSERT INTO isbn (isbn, year, publisher, author, title)"
    #    "VALUES (%s, %s, %s, %s, %s)"
    #    )
    #    data = (isbn, year, publisher, author, title)
    #    #cursor = connection.cursor()
    #    cursor.execute(mySql_insert_query, data)
    #    connection.commit()
    #    print(cursor.rowcount, "Record successfully inserted into isbn")
        
            
            
            #print ("Author(s):\t", author)
            #print ("Title:\t\t", title)
            #print ("ISBN:\t\t", isbn13)
            #print ("Year:\t\t", year)
            #print ("Publisher:\t", publisher)




def exportDb():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    
    ============================================================================
    """

    

    for j in good_dict:

        #j = 1

        print(good_dict.get('isbn'))
        #year = good_dict.get('year'[j])
        #publisher = good_dict.get('publisher'[j])
        #author = good_dict.get('author'[j])
        #title = good_dict.get('title'[j])
        
        
        
        
        #data = (isbn, year, publisher, author, title)

        #mySql_insert_query = (
        #"INSERT INTO isbn (isbn, year, publisher, author, title)"
        #"VALUES (%s, %s, %s, %s, %s)"
        #)
        ##data = (isbn, year, publisher, author, title)
        ##cursor = connection.cursor()
        #cursor.execute(mySql_insert_query, data)
        #connection.commit()
        #print(cursor.rowcount, "Record successfully inserted into isbn")

        #print(data)

        #print(isbn)

        
        
        #j = j + 1
        #cursor.close()            

def main():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """
    #dbConnect()
    createLists()
    getInfo()
    #exportBad()
    #exportDb()
    #print(good_list)
    #print(good_dict)
    

if __name__ == "__main__":
    main()
