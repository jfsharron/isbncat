"""
================================================================================
 Program:           newMain.py
 Software Engineer: Jonas Sharron
 Date:              01-October-2022

 Purpose:   This program will process isbn's stored in a file and export them to 
            a MySQL database.  The program will also check for isbn's that are
            not represented in the search service and export them to list for
            evaluation by the user.  A manuel entry and editing method is also 
            provided for user interaction.

Related Files:      
            variable            file                purpose
            --------------------------------------------------------------------
            workbookName        inventory.xlsx      -data input file
            dataframeName       dataframe.xlsx      -dataframe created to remove
                                                     duplicate isbn's from data
                                                     input file
            dbIsbnxls           dbxls.xlsx          -dataframe created from MySQL 
                                                     to create dbIsbn list 

Lists:
            isbn_list[]         -list created from input file (workbookName), 
                                 duplicates removed
            
            bad_list and good_list are created from isbn_list
            --------------------------------------------------------------------
            bad_list[]          -list of isbn's not available in search service
            good_list[]         -list of isbn's available in search service

            
            dbIsbn_list[]       -list generated from records in MySQL database
            dup_list[]          -list of duplicate values found in both 
                                 dbIsbn_list and isbn_list (these values are 
                                 removed from the isbn_list) 

        * these lists are written to text files in the log directory 

******************************************************************************** 
   TO PROPERLY START THE PROGRAM, YOU MUST PROVIDE A COMMAND LINE ARGUMENT 
   SPECIFYING THE LOCATION AND NAME OF THE FILE CONTAINING YOUR CREDENTIALS
   FOR EXAMPLE:
      python newMain.py C:\\Users\\user1\\Documents\\isbn_info
      * only use single backslashes (\) in the path
********************************************************************************

================================================================================
"""

import sys
import pandas as pd
from isbnlib import meta
from isbnlib.registry import bibformatters
import openpyxl
import mysql.connector
from mysql.connector import Error
import functools
import xlsxwriter
import pandas.io.sql as sql
import numpy as np
import pickle
from numpy import loadtxt
from prettytable import from_db_cursor
from prettytable import PrettyTable
import datetime
from datetime import date
import os
from termcolor import colored, cprint 
from colorama import Fore, Back, Style 
from tabulate import tabulate
#from win32printing import Printer
import fpdf
import colorama
from colorama import Fore, Back, Style
import getopt
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch
import webbrowser
import time
from pretty_html_table import build_table
from sqlalchemy import create_engine

# retrieve stored information
# ===========================
filename = "dump"
file = open(filename, 'rb')
new = pickle.load(file)
lines = loadtxt(new, dtype=str, comments="#", delimiter=",", unpack=False)
file.close()

USER        = str(lines[0])
XWORD       = str(lines[1])
HOST        = str(lines[2])
DATABASE    = str(lines[3])
DATAFILE    = str(lines[4])
FNAME       = str(sys.argv[1])
SERVICE     = "openl"
VERNO       = "1.0.0"
VERNA       = "Trinidad"



# define external files
# ======================
workbookName    = DATAFILE
dataframeName   = "dataframe.xlsx"
dbIsbnxls       = "dbxls.xlsx"

# initialize lists
# =================
bad_list        = []
good_list       = []
isbn_list       = []
dbIsbn_list     = []
dup_list        = []

# establish database connection
# =============================
try:
    SUSERNAME     = str(lines[0])
    SPASSWORD     = str(lines[1])
    SHOST         = str(lines[2])
    SDATABASE     = str(lines[3])

    CONNECTION = mysql.connector.connect(user=SUSERNAME, password=SPASSWORD,
    host=SHOST, database=SDATABASE)
    if CONNECTION.is_connected():
        db_Info = CONNECTION.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = CONNECTION.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

# ==============================================================================
# user functions
# ==============================================================================

def createLists():
    """
    ============================================================================
    Function:       createLists()
    Purpose:        filters input data into two lists (good_list and bad_list)
                    depending on the availability of data in the search service
                    (filtered by return data from ('Authors') field)
    Parameter(s):   -None- (processes data in external file)
    Return:         -None- (propagates data in good_list and bad_list)
    ============================================================================
    """
    # iterate through data file and for existence in search service
    print("Checking search service for data file information . . .")
    for i in isbn_list:
        isbn = i

        SERVICE = "openl"
        bibtex = bibformatters["bibtex"]
        
        # if data is not available in search service, add isbn to bad_list
        # otherwise add isbn to good_list
        # ==================================================================
        meta_dict = meta(isbn, service='default')
        if meta_dict.get('Authors') is None:
            bad_list.append(isbn)
        else:
            good_list.append(isbn)  

    # data file check completion message
    print("Check for datafile information completed, data with available "
        "information exported to good_list, data without information exported"
        " to bad_list . . .")

def getInfo():    
    """
    ============================================================================
    Function:       getInfo()
    Purpose:        retrieves information (author, title, isbn, year, publisher)
                    from search service and propagates MySQL database with hew
                    record.
    Parameter(s):   -None- (processes data in good_list)
    Return:         -None- (exports data to MySQL database)
    ============================================================================
    """
    global CONNECTION
    global SERVICE

    print("Connecting to search service . . .")
    print("Retrieving information for good_list . . .")
    for i in good_list:
            isbn = i
            #SERVICE = "openl"
            bibtex = bibformatters["bibtex"]
            meta_dict = meta(isbn, service='default')
            aut = str(meta_dict['Authors'])
            aut = aut.replace("[","")
            aut = aut.replace("]","")
            author = aut.replace("'","")
            title = meta_dict['Title']
            isbn = meta_dict['ISBN-13']
            year = meta_dict['Year']
            publisher = meta_dict['Publisher']

            # define data to retrieve
            data = (isbn, year, publisher, author, title)

            # SQL query to insert data into to db
            mySql_insert_query = (
            "INSERT INTO isbn (isbn, year, publisher, author, title)"
            "VALUES (%s, %s, %s, %s, %s)"
            )
        
            cursor = CONNECTION.cursor()
            cursor.execute(mySql_insert_query, data)
            CONNECTION.commit()

    # completion message
    print("Available information added to SQL database")            

def exportLists():
    """
    ============================================================================
    Function:       exportLists()
    Purpose:        exports lists external log files then clears them
    Parameter(s):   -None- (processes from lists)
    Return:         -None- (generates external log files)    
    ============================================================================
    """
    # create datetime string for log entries
    # ---------------------------------------
    dt = datetime.datetime.now()
    dt = str(dt)

    # write lists to log files
    # ------------------------
    with open('log/bad_list', 'a') as fp1:
        for i in bad_list:
            fp1.write(dt + "\t\t + %s\n" % i)    
    fp1.close()
    bad_list.clear()

    with open('log/good_list', 'a') as fp2:
        for i in good_list:
            fp2.write(dt + "\t\t + %s\n" % i)    
    fp2.close()
    good_list.clear()

    with open('log/isbn_list', 'a') as fp3:
        for i in isbn_list:
            fp3.write(dt + "\t\t + %s\n" % i)    
    fp3.close()
    isbn_list.clear()

    with open('log/dbIsbn_list', 'a') as fp4:
        for i in dbIsbn_list:
            fp4.write(dt + "\t\t + %s\n" % i)    
    fp4.close()
    dbIsbn_list.clear()

    with open('log/dup_list', 'a') as fp5:
        for i in dup_list:
            fp5.write(dt + "\t\t + %s\n" % i)    
    fp5.close()
    dup_list.clear()

def preProcess():
    """
    ============================================================================
    Function:       preProcess()
    Purpose:        looks for and removes duplicate isbn's from import file and 
                    pre-existing db 
    Parameter(s):   -None- (reads data from input file and db)
    Return:         -None- (writes data (duplicates removed) to isbn_list)
    ============================================================================
    """
    print("Checking for duplicates in source file . . .")


    # remove duplicates from isbn spreadsheet, save in dataframe spreadsheet
    data = pd.read_excel(workbookName, sheet_name = 'data', usecols = ['isbn'])
    data_first_record = data.drop_duplicates(keep="first")

    writer = pd.ExcelWriter(dataframeName, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    data_first_record.to_excel(writer, sheet_name='Sheet1')

    # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # Add numeric cell formats.
    format1 = workbook.add_format({'num_format': '###0'})

    worksheet.set_column(1, 1, 18, format1)

    writer.save()

    dframe = openpyxl.load_workbook(dataframeName)
    sh = dframe.active

    # send dataframe data to isbn_list
    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            isbn = str(cell.value)
            isbn_list.append(isbn)
    
    # create dbIsbn_list from database
    print("Checking for duplicates in database file . . .")
    query = "SELECT isbn FROM isbn"
    df = sql.read_sql('SELECT isbn FROM isbn', CONNECTION)
    df.to_excel(dbIsbnxls)

    dbexcel = openpyxl.load_workbook(dbIsbnxls)
    sh = dbexcel.active

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            isbn = str(cell.value)
            dbIsbn_list.append(isbn)
    
    # compare isbn_list and dbIsbn_list and create intersection (duplicates) list
    a = (isbn_list)
    b = (dbIsbn_list)


    intersection = set(a).intersection(b)
    

    # remove intersection (duplicates) list values from isbn_list
    for value in intersection:
        if value in isbn_list:
            dup_list.append(value)
            isbn_list.remove(value)

    # completion of duplicate check message
    print("Duplicate check completed, duplicates removed and exported to "
        "dup_list . . .")

def getGenre():
    """
    ============================================================================
    Function:       getGenre()
    Purpose:        imports genre information into MySQL database from external 
                    file
    Parameter(s):   -None- (reads data from input file and db)
    Return:         -None- (writes genre data to MySQL)
    ============================================================================
    """    
       
    # create datafrane from external file
    print('Opening datafile . . .')
    gframe = openpyxl.load_workbook(workbookName)
    data = gframe.active
    
    # define max row with x and y variables
    x = data.max_row
    y = 'B' + str(x) 
    
    cells = data['A2' : y]

    # iterate through external file to retrieve genre values
    print('Reading datafile values . . .')
    print('Inserting values into database . . .')
    for c1, c2 in cells:
        gisbn = (c1.value)
        gisbn = str(gisbn)
        genre = (c2.value)
        genre = str(genre)

        sqldata = (genre, gisbn)

        # define MySQL query and import genre values into database
        genre_query = ("UPDATE isbn SET genre = (%s) WHERE isbn = (%s)")
        
        cursor = CONNECTION.cursor()
        cursor.execute(genre_query, sqldata)
        CONNECTION.commit()

def menu():
    """
    ============================================================================
    Function:       menu()
    Purpose:        entry point to allow user interaction with program
    Parameter(s):   -None-
    Return:         users desired action
    ============================================================================
    """
    os.system('cls')
    # main menu
    #----------
    goAgain = 1

    while goAgain == 1:
        # format screen
        # --------------
        now = datetime.datetime.now()
        print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
        print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
        print("-----------------------".rjust(80))
        print(Style.RESET_ALL)
        print('')
        print(Fore.GREEN + 'MAIN MENU')
        print(Fore.GREEN + '-------------------')
        print(Style.RESET_ALL)
        print('1\tSet System Parameters')
        print('2\tProgram Functions')
        print('3\tReports')
        print('')
        print('')
        print('')
        print(Fore.RED + '0\tEXIT')
        print(Style.RESET_ALL)
        print('')
        print('')    

        # menu options based on user input
        # ---------------------------------
        menuOption = input("selection: ")

        if menuOption == '1':
            sysParmMenu()
        elif menuOption == '2':
            programFunctMenu()
        elif menuOption == '3':
            newReport() 
        elif menuOption == '0':    
            goAgain = 0 

        os.system('cls')   

def sysParmMenu():
    """
    ============================================================================
    Function:       sysParmMenu()
    Purpose:        provides user options for editing program parameters
    Parameter(s):   -None- 
    Return:         users desired action
    ============================================================================
    """
    goAgain = 1
 
    # global values for MySQL
    # -----------------------
    global USER 
    global XWORD
    global HOST
    global DATABASE
    global DATAFILE
    global FNAME
    
    # system parameters menu
    # ----------------------
    while goAgain == 1:
        os.system('cls')
        now = datetime.datetime.now()
        print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
        print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
        print("-----------------------".rjust(80))
        print(Style.RESET_ALL)
        print('')
        print(Fore.GREEN + 'SET SYSTEM PARAMETERS')
        print(Fore.GREEN + '-------------------')
        print(Style.RESET_ALL)
        print(Fore.RED +"PLEASE NOTE: if the values filename, location, or values "
            "are edited, the program will have to be restarted for the changes to "
            "take effect")
        print(Style.RESET_ALL)
        print('')
        print('1\tDisplay Values File Filepath and Name')
        print('2\tDisplay and Edit Values File')
        print('')
        print('')
        print('')
        print(Fore.RED + '0\tRETURN')
        print(Style.RESET_ALL)
        print('')
        print('')    

        # menu options based on user input
        # ---------------------------------
        menuOption = input("selection: ")

        if menuOption == '1':                     
            print('')
            print(Fore.YELLOW + str(FNAME)) 
            print(Style.RESET_ALL)
            print('')
            wait = input("Press ENTER to return")
        elif menuOption == '2':
            goAgain2 = 1
            while goAgain2 == 1:
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print('')
                print(Fore.GREEN + 'VALUES FILE')
                print(Fore.GREEN + '-------------------')
                print(Style.RESET_ALL)
                print(Fore.RED + "These values should be kept secure and not "
                     "shared")
                print(Style.RESET_ALL)
                print('')
                # submenu for MySQL options
                # -------------------------
                print('1\tSet MySQL username')
                print('2\tDisplay MySQL username')
                print('3\tMySQL password')
                print('4\tDisplay MySQL password')
                print('5\tSet MySQL host address')
                print('6\tDisplay MySQL host address')
                print('7\tSet MySQL database name')
                print('8\tDisplay MySQL database name')
                print('9\tSet data file name and location')
                print('10\tDisplay data file name and location')
                print('')
                print('')
                print('')
                print(Fore.RED + '0\tRETURN')
                print(Style.RESET_ALL)
                print('')
                print('')    

                # menu options based on user input
                # ---------------------------------
                submenuOption = input("selection: ")
                if submenuOption == '1':
                    os.system('cls')
                    now = datetime.datetime.now()
                    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                    print("-----------------------".rjust(80))
                    print(Style.RESET_ALL)
                    print("Current MySQL username is: " + USER)
                    tempUser = input("Please Enter your new MySQL username: ")
                    print('')
                    print(Fore.YELLOW + 
                        "You are about to reset to MySQL user name to " + tempUser
                         + " are you sure?") 
                    print(Style.RESET_ALL)
                    print('')
                    resetUser = input("Type YES to make change: ")
                    if resetUser == 'YES':
                        USER = tempUser
                        print('MySQL username changed')
                        print('')
                        wait = input("Press ENTER to return")     
                    else:
                        print('MySQL username not changed')
                        print('')
                        wait = input("Press ENTER to return")                    
                elif submenuOption == '2':
                    print('')
                    print(Fore.YELLOW + USER) 
                    print(Style.RESET_ALL)
                    print('')
                    wait = input("Press ENTER to return")
                elif submenuOption == '3':
                    os.system('cls')
                    now = datetime.datetime.now()
                    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                    print("-----------------------".rjust(80))
                    print(Style.RESET_ALL)
                    print("Current MySQL password is: " + XWORD)
                    tempXword = input("Please Enter your new MySQL password: ")
                    print('')
                    print(Fore.YELLOW + 
                        "You are about to reset to MySQL password to " + tempXword
                         + " are you sure?") 
                    print(Style.RESET_ALL)
                    print('')
                    resetXword = input("Type YES to make change: ")
                    if resetXword == 'YES':
                        XWORD = tempXword
                        print('MySQL password changed')
                        print('')
                        wait = input("Press ENTER to return")     
                    else:
                        print('MySQL password not changed')
                        print('')
                        wait = input("Press ENTER to return")
                elif submenuOption == '4':
                    print('')
                    print(Fore.YELLOW + XWORD) 
                    print(Style.RESET_ALL)
                    print('')
                    wait = input("Press ENTER to return")
                elif submenuOption == '5':
                    os.system('cls')
                    now = datetime.datetime.now()
                    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                    print("-----------------------".rjust(80))
                    print(Style.RESET_ALL)
                    print("Current MySQL host is: " + HOST)
                    tempHost = input("Please Enter your new MySQL host: ")
                    print('')
                    print(Fore.YELLOW + 
                        "You are about to reset to MySQL host to " + tempHost
                         + " are you sure?") 
                    print(Style.RESET_ALL)
                    print('')
                    resetHost = input("Type YES to make change: ")
                    if resetHost == 'YES':
                        HOST = tempHost
                        print('MySQL host changed')
                        print('')
                        wait = input("Press ENTER to return")     
                    else:
                        print('MySQL host not changed')
                        print('')
                        wait = input("Press ENTER to return")
                elif submenuOption == '6':
                    print('')
                    print(Fore.YELLOW + HOST) 
                    print(Style.RESET_ALL)
                    print('')
                    wait = input("Press ENTER to return")
                elif submenuOption == '7':
                    os.system('cls')
                    now = datetime.datetime.now()
                    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                    print("-----------------------".rjust(80))
                    print(Style.RESET_ALL)
                    print("Current MySQL database is: " + DATABASE)
                    tempDb = input("Please Enter your new MySQL database: ")
                    print('')
                    print(Fore.YELLOW + 
                        "You are about to reset to MySQL database to " + tempDb
                         + " are you sure?") 
                    print(Style.RESET_ALL)
                    print('')
                    resetDb = input("Type YES to make change: ")
                    if resetDb == 'YES':
                        DATABASE = tempDb
                        print('MySQL database changed')
                        print('')
                        wait = input("Press ENTER to return")     
                    else:
                        print('MySQL database not changed')
                        print('')
                        wait = input("Press ENTER to return")
                elif submenuOption == '8':
                    print('')
                    print(Fore.YELLOW + DATABASE) 
                    print(Style.RESET_ALL)
                    print('')
                    wait = input("Press ENTER to return")
                elif submenuOption == '9':
                    os.system('cls')
                    now = datetime.datetime.now()
                    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                    print("-----------------------".rjust(80))
                    print(Style.RESET_ALL)
                    print(Fore.YELLOW + "Please use double backslashes (\\\) when "
                    "defining file path") 
                    print(Style.RESET_ALL)
                    print("Current MySQL datafile is: " + DATAFILE)
                    tempDf = input("Please Enter your new MySQL datafile: ")
                    print('')
                    print(Fore.YELLOW + 
                        "You are about to reset to MySQL datafile to " + tempDf
                         + " are you sure?") 
                    print(Style.RESET_ALL)
                    print('')
                    resetDf = input("Type YES to make change: ")
                    if resetDf == 'YES':
                        DATAFILE = tempDf
                        print('MySQL datafile changed')
                        print('')
                        wait = input("Press ENTER to return")     
                    else:
                        print('MySQL datafile not changed')
                        print('')
                        wait = input("Press ENTER to return")
                elif submenuOption == '10':
                    print('')
                    print(Fore.YELLOW + DATAFILE) 
                    print(Style.RESET_ALL)
                    print('')
                    wait = input("Press ENTER to return")
                elif submenuOption == '0':
                    goAgain2 = 0
        # elif statement to return to rewrite pickle file and return to previous
        # menu
        # ----------------------------------------------------------------------
        elif menuOption == '0':
            value = str(FNAME)
            vvfile = (USER+","+XWORD+","+HOST+","+DATABASE+","+DATAFILE)
            vfile = open(FNAME, "w")
            n = vfile.write(vvfile)
            vfile.close()
            filename = "dump"
            file = open(filename, 'wb')
            pickle.dump(value, file)
            file.close()
            goAgain = 0    

def programFunctMenu():
    """
    ============================================================================
    Function:       programFunctMenu()
    Purpose:        provides user options for performing program functions
    Parameter(s):   -None- 
    Return:         users desired action
    ============================================================================
    """
    global CONNECTION
    os.system('cls')

    now = datetime.datetime.now()
    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
    print("-----------------------".rjust(80))
    print(Style.RESET_ALL)

    goAgain = 1

    # display main Program Functions menu and select options
    # -------------------------------------------------------
    while goAgain == 1:
        print('')
        print(Fore.GREEN + 'PROGRAM FUNCTIONS')
        print(Fore.GREEN + '-------------------')
        print(Style.RESET_ALL)
        print('1\tSearch for a Record')
        print('2\tEdit a Record')
        print('3\tManually Add a Record')
        print('4\tDelete a Record')
        print('')
        print('')
        print('5\tImport Records (with genre)')
        print('6\tImport Records (without genre)')
        print('')
        print('')
        print('7\tImport Genre')
        print('')
        print('')
        print('')
        print(Fore.RED + '0\tRETURN')
        print(Style.RESET_ALL)
        print('')
        print('')    

        menuOption = input("selection: ")

        # select program function options
        # -------------------------------
        if menuOption == '1':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print('')
            qisbn = input("Enter ISBN to search for: ")
            qisbn =str(qisbn)
            mysql_search_query = ("SELECT * FROM isbn WHERE isbn = " + qisbn)
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(mysql_search_query)    
            mytable = from_db_cursor(cursor)
            mytable.align = "l"
            print(mytable)
            print('')
            wait = input("Press ENTER to return")
        if menuOption == '2':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print('')
            qisbn = input("Enter ISBN to edit: ")
            qisbn =str(qisbn)
            mysql_search_query = ("SELECT * FROM isbn WHERE isbn = " + qisbn)
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(mysql_search_query)
            for row in cursor:
                eID = str(row[0])
                print("ID:\t\t" + eID)
                eIsbn = row[1]
                print("ISBN:\t\t" + eIsbn)
                eYear = row[2]
                print("Year:\t\t" + eYear)
                ePublisher = row[3]
                print("Publisher:\t" + ePublisher)
                eAuthor = row[4]
                print("Author:\t\t" + eAuthor)
                eTitle = row[5]
                print("Title:\t\t" + eTitle)
                eGenre = row[6]
                print("Genre:\t\t" + eGenre)
                print('')
            # edit isbn options sub menu and options
            # ---------------------------------------
            print('1\tEdit ISBN')
            print('2\tEdit Year')
            print('3\tEdit Publisher')
            print('4\tEdit Author')
            print('5\tEdit Title')
            print('6\tEdit Genre')
            print('')
            print(Fore.RED + '0\tRETURN')
            print(Style.RESET_ALL)
            print('')
            editOption = input("selection: ")
            if editOption == '1':
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print("Current ISBN is: " + eIsbn)
                tempIsbn = input("Please Enter your new ISBN: ")
                print('')
                print(Fore.YELLOW + 
                    "You are about to change the ISBN to " + tempIsbn
                     + " are you sure?") 
                print(Style.RESET_ALL)
                print('')
                resetIsbn = input("Type YES to make change: ")
                if resetIsbn == 'YES':
                    eIsbn = tempIsbn
                    print('ISBN changed')
                    print('')
                    mysql_change_query = ("UPDATE isbn SET isbn = " + eIsbn +
                                        " WHERE isbn_id = " + eID)
                    cursor = CONNECTION.cursor()
                    cursor.execute(mysql_change_query)
                    CONNECTION.commit()
                    wait = input("Press ENTER to return")    
                else:
                    print('ISBN not changed')
                    print('')
                    wait = input("Press ENTER to return")
            if editOption == '2':
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print("Current Year is: " + eYear)
                tempYear = input("Please Enter your new Year: ")
                print('')
                print(Fore.YELLOW + 
                    "You are about to change the Year to " + tempYear
                     + " are you sure?") 
                print(Style.RESET_ALL)
                print('')
                resetYear = input("Type YES to make change: ")
                if resetYear == 'YES':
                    eYEAR = tempYear
                    print('Year changed')
                    print('')
                    mysql_change_query = ("UPDATE isbn SET year = " + eYear +
                                        " WHERE isbn_id = " + eID)
                    cursor = CONNECTION.cursor()
                    cursor.execute(mysql_change_query)
                    CONNECTION.commit()
                    wait = input("Press ENTER to return")    
                else:
                    print('Year not changed')
                    print('')
                    wait = input("Press ENTER to return")
            if editOption == '3':
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print("Current Publisher is: " + ePublisher)
                tempPublisher = input("Please Enter your new Publisher: ")
                print('')
                print(Fore.YELLOW + 
                    "You are about to change the Publisher to " + tempPublisher
                     + " are you sure?") 
                print(Style.RESET_ALL)
                print('')
                resetPublisher = input("Type YES to make change: ")
                if resetPublisher == 'YES':
                    ePublisher = tempPublisher
                    print('Publisher changed')
                    print('')
                    mysql_change_query = ("UPDATE isbn SET publisher = " + 
                                         ePublisher + " WHERE isbn_id = " + eID)
                    cursor = CONNECTION.cursor()
                    cursor.execute(mysql_change_query)
                    CONNECTION.commit()
                    wait = input("Press ENTER to return")
                else:
                    print('Publisher not changed')
                    print('')
                    wait = input("Press ENTER to return")
            if editOption == '4':
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print("Current Author is: " + eAuthor)
                tempAuthor = input("Please Enter your new Author: ")
                print('')
                print(Fore.YELLOW + 
                    "You are about to change the Author to " + tempAuthor
                     + " are you sure?") 
                print(Style.RESET_ALL)
                print('')
                resetAuthor = input("Type YES to make change: ")
                if resetAuthor == 'YES':
                    eAuthor = tempAuthor
                    print('Author changed')
                    print('')
                    mysql_change_query = ("UPDATE isbn SET author = " + eAuthor+
                                        " WHERE isbn_id = " + eID)
                    cursor = CONNECTION.cursor()
                    cursor.execute(mysql_change_query)
                    CONNECTION.commit()
                    wait = input("Press ENTER to return")    
                else:
                    print('Author not changed')
                    print('')
                    wait = input("Press ENTER to return") 
            if editOption == '5':
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print("Current Title is: " + eTitle)
                tempTitle = input("Please Enter your new Title: ")
                print('')
                print(Fore.YELLOW + 
                    "You are about to change the Title to " + tempTitle
                     + " are you sure?") 
                print(Style.RESET_ALL)
                print('')
                resetTitle= input("Type YES to make change: ")
                if resetTitle == 'YES':
                    eTitle= tempTitle
                    print('Title changed')
                    print('')
                    mysql_change_query = ("UPDATE isbn SET title = " + eTitle +
                                        " WHERE isbn_id = " + eID)
                    cursor = CONNECTION.cursor()
                    cursor.execute(mysql_change_query)
                    CONNECTION.commit()
                    wait = input("Press ENTER to return")    
                else:
                    print('Title not changed')
                    print('')
                    wait = input("Press ENTER to return")
            if editOption == '6':
                os.system('cls')
                now = datetime.datetime.now()
                print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
                print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
                print("-----------------------".rjust(80))
                print(Style.RESET_ALL)
                print("Current Genre is: " + eGenre)
                tempGenre = input("Please Enter your new Genre: ")
                print('')
                print(Fore.YELLOW + 
                    "You are about to change the Genre to " + tempGenre 
                     + " are you sure?") 
                print(Style.RESET_ALL)
                print('')
                resetGenre  = input("Type YES to make change: ")
                if resetGenre == 'YES':
                    eGenre = tempGenre 
                    print('Genre changed')
                    print('')
                    mysql_change_query = ("UPDATE isbn SET genre = " + eGenre +
                                        " WHERE isbn_id = " + eID)
                    cursor = CONNECTION.cursor()
                    cursor.execute(mysql_change_query)
                    CONNECTION.commit()
                    wait = input("Press ENTER to return")    
                else:
                    print('Genre not changed')
                    print('')
                    wait = input("Press ENTER to return")    
                    # -----------------------------------
                    # end isbn edit options                             
        if menuOption == '3':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print('')
            print('')
            print(Fore.YELLOW + "Please provide a "
                    "13 digit ISBN (\x1B[3mrequired\x1B[0m" + Fore.YELLOW + "), "
                    "published year (\x1B[3moptional\x1B[0m" + Fore.YELLOW + "), "
                    "publisher (\x1B[3moptional\x1B[0m" + Fore.YELLOW + "), "
                    "author (\x1B[3moptional\x1B[0m" + Fore.YELLOW + "), "
                    "title (\x1B[3mrequired\x1B[0m" + Fore.YELLOW + "), "
                    "and genre (\x1B[3optional\x1B[0m" + Fore.YELLOW + ")")
            print('')
            print("Please enter year in 4-digit format")
            print(Style.RESET_ALL)
            print('')
            print('')
            aISBN = input("Enter ISBN:\t\t ")
            aYear = input("Enter Year:\t\t ")
            aPublisher = input("Enter Publisher:\t ")
            aAuthor = input("Enter Author(s):\t ")
            aTitle = input("Enter Title:\t\t ")
            aGenre = input("Enter Genre:\t\t ")
            print('')
            data = (aISBN, aYear, aPublisher, aAuthor, aTitle)
            mySql_insert_query = (
            "INSERT INTO isbn (isbn, year, publisher, author, title)"
            "VALUES (%s, %s, %s, %s, %s)"
            )
            cursor = CONNECTION.cursor()
            cursor.execute(mySql_insert_query, data)
            CONNECTION.commit()
            print(Fore.YELLOW + "Record added")
            print(Style.RESET_ALL)
            print('')
            mysql_search_query = ("SELECT * FROM isbn WHERE isbn = " + aISBN)
            cursor.execute(mysql_search_query)    
            mytable = from_db_cursor(cursor)
            mytable.align = "l"
            print(mytable)
            print('')
            wait = input("Press ENTER to return")
        if menuOption == '4':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print('')
            print('')
            dIsbn = input('Enter isbn to delete: ')
            print(dIsbn)
            print('')
            mysql_search_query = ("SELECT * FROM isbn WHERE isbn = " + dIsbn)
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(mysql_search_query)    
            mytable = from_db_cursor(cursor)
            mytable.align = "l"
            print(mytable)
            print('')
            print(Fore.YELLOW + 
                   "You are about to delete the ISBN " + dIsbn
                    + " are you sure?") 
            print(Style.RESET_ALL)
            print('')
            deleteIsbn = input("Type YES to make change: ")
            if deleteIsbn == 'YES':
                print('ISBN deleted')
                print('')
                mysql_delete_query = ("DELETE FROM isbn WHERE isbn = " + dIsbn)
                cursor = CONNECTION.cursor()
                cursor.execute(mysql_delete_query)
                CONNECTION.commit()
                wait = input("Press ENTER to return")    
            else:
                print('ISBN not deleted')
                print('')
                wait = input("Press ENTER to return")
        if menuOption == '5':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print(Fore.YELLOW + "You are about to import records (including genre)"
                "into your MySQL database")
            print('')
            print(Style.RESET_ALL)
            import2 = input("Type YES to proceed: ")
            if import2 == 'YES':
                print('')
                print('import being processed . . .')
                print('')
                searchService()
                preProcess()
                createLists()
                getInfo()
                getGenre()   
                exportLists()
                wait = input("Press ENTER to return")    
            else:
                print('import not processed')
                print('')
                wait = input("Press ENTER to return")
        if menuOption == '6':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print(Fore.YELLOW + "You are about to import records (without genre)"
                "into your MySQL database")
            print('')
            print(Style.RESET_ALL)
            import2 = input("Type YES to proceed: ")
            if import2 == 'YES':
                print('')
                print('import being processed . . .')
                print('')
                searchService()
                preProcess()
                createLists()
                getInfo()  
                exportLists()
                wait = input("Press ENTER to return")    
            else:
                print('import not processed')
                print('')
                wait = input("Press ENTER to return")  
        if menuOption == '7':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print(Fore.YELLOW + "You are about to import genre values into your "
                "MySQL database")
            print('')
            print(Style.RESET_ALL)
            import3 = input("Type YES to proceed: ")
            if import3 == 'YES':
                print('')
                print('import being processed . . .')
                print('')
                getGenre()
                exportLists()
                wait = input("Press ENTER to return")    
            else:
                print('import not processed')
                print('')
                wait = input("Press ENTER to return")
        elif menuOption == '0':
            goAgain = 0

def newReport():
    """
    ============================================================================
    Function:       newReport()
    Purpose:        provides user options for accessing reports (v2)
    Parameter(s):   -None- 
    Return:         users desired action
    ============================================================================
    """     
    os.system('cls')

    now = datetime.datetime.now()
    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
    print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
    print("-----------------------".rjust(80))
    print(Style.RESET_ALL)

    goAgain = 1

    # reports menu and options
    # ------------------------
    while goAgain == 1:
        print('')
        print(Fore.GREEN + 'REPORTS')
        print(Fore.GREEN + '-------------------')
        print(Style.RESET_ALL)
        print('1\tShow Report Index')
        print('2\tFilter Report Index')
        print('3\tShow Report By Name')        
        print('')
        print('4\tCreate New Report')
        print('')
        print(Fore.RED + '0\tRETURN')
        print(Style.RESET_ALL)
        print('')
        print('')    

        menuOption = input("selection: ")

        # show report index
        # -------------------
        if menuOption == '1':
            print(Fore.GREEN + 'Report Index')
            print(Fore.GREEN + '-------------------')
            print(Style.RESET_ALL)
            print('')     
            mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                    cq_desc AS \"report description\", \
                                    cqType_name AS \"report type\", \
                                    cq_parameters AS \"parameters(0=no, 1=yes)\",\
                                    cq_creator AS \"report orginator\", \
                                    cq_created AS \"origination date\" \
                                    FROM cQuery p INNER JOIN cqType c \
                                    ON p.cq_type = c.cqType_value")
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(mysql_search_query)    
            mytable = from_db_cursor(cursor)
            mytable.align = "l"
            print(mytable)
            print('')
            # send report to browser
            # -----------------------
            printRep = input(Fore.YELLOW + 'To send this report to the browser '
                            'for printing or saving enter b or B, otherwise press '
                            'enter to return: ')
            print(Style.RESET_ALL)
            if printRep == "b" or printRep == "B":
                # generate data for report
                # ------------------------
                mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                    cq_desc AS \"report description\", \
                                    cqType_name AS \"report type\", \
                                    cq_parameters AS \"parameters(0=no, 1=yes)\",\
                                    cq_creator AS \"report orginator\", \
                                    cq_created AS \"origination date\" \
                                    FROM cQuery p INNER JOIN cqType c \
                                    ON p.cq_type = c.cqType_value")
                cursor = CONNECTION.cursor(buffered = True)
                cursor.execute(mysql_search_query)
                mytable = pd.read_sql("SELECT cq_name AS \"report name\", \
                                    cq_desc AS \"report description\", \
                                    cqType_name AS \"report type\", \
                                    cq_parameters AS \"parameters(0=no, 1=yes)\",\
                                    cq_creator AS \"report orginator\", \
                                    cq_created AS \"origination date\" \
                                    FROM cQuery p INNER JOIN cqType c \
                                    ON p.cq_type = c.cqType_value", CONNECTION)
                pd.set_option('display.expand_frame_repr', False)
                mytable2 = build_table(mytable,
                                     'grey_light',
                                     font_size = 'small',
                                     font_family = 'Open Sans, courier',
                                     text_align = 'left ')
                # generate html content
                # ---------------------
                html_content = f"<html> \
                                <head> <h2> ISBN22 Report Index - All Records\
                                </h2> \
                                <h3> <script>\
                                var timestamp = Date.now();\
                                var d = new Date(timestamp);\
                                document.write(d);\
                                </script>\
                                </h3>\
                                </head> \
                                <body> {mytable2} \
                                </body> \
                                </html>"
                with open("report/report_index.html", "w") as html_file:
                    html_file.write(html_content)
                    print("Created")
                time.sleep(2)
                # display in browser
                # ------------------
                webbrowser.open_new_tab("report\\report_index.html")
                print('')
                wait = input("Press ENTER to return")         
        # display filtered index
        # ======================
        elif menuOption == '2':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            print("Available filters are:")
            print("----------------------")
            print('')
            mysql_search_query = ("SELECT cqType_value AS 'value', \
                                cqType_name AS 'name', \
                                cqType_desc AS 'description' FROM cqType")
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(mysql_search_query)    
            mytable = from_db_cursor(cursor)
            mytable.align = "l"
            print(mytable)
            print('')
            print('')  
            filt = input("Please select a filter value to display reports by: ")
            # filter by type "test"
            #----------------------
            if filt == "1":
                cqn = '"test"'
                print(Fore.GREEN + 'Filtered Report Index')
                print(Fore.GREEN + '----------------------')
                print(Style.RESET_ALL)
                print('')     
                mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value\
                                        WHERE cqType_name = " + cqn)
                cursor = CONNECTION.cursor(buffered = True)
                cursor.execute(mysql_search_query)    
                mytable = from_db_cursor(cursor)
                mytable.align = "l"
                print(mytable)
                print('')
                # send report to browser
                # -----------------------
                printRep = input(Fore.YELLOW + 
                                'To send this report to the browser for printing'
                                ' or saving enter b or B, otherwise press '
                                'enter to return: ')
                print(Style.RESET_ALL)
                if printRep == "b" or printRep == "B":
                    # generate data for report
                    # ------------------------
                    mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value WHERE \
                                        cqType_name = " + cqn)
                    cursor = CONNECTION.cursor(buffered = True)
                    cursor.execute(mysql_search_query)
                    mytable = pd.read_sql("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value WHERE \
                                        cqType_name = " + cqn, CONNECTION)
                    pd.set_option('display.expand_frame_repr', False)
                    mytable2 = build_table(mytable,
                                         'grey_light',
                                         font_size = 'small',
                                         font_family = 'Open Sans, courier',
                                         text_align = 'left ')
                    # generate html content
                    # ---------------------
                    html_content = f"<html> \
                                    <head> <h2> ISBN22 Filtered Report Index \
                                    <br>Filter: {cqn} \
                                    </h2> \
                                    <h3> <script>\
                                    var timestamp = Date.now();\
                                    var d = new Date(timestamp);\
                                    document.write(d);\
                                    </script>\
                                    </h3>\
                                    </head> \
                                    <body> {mytable2} \
                                    </body> \
                                    </html>"
                    with open("report/filt_report_index.html", "w") as html_file:
                        html_file.write(html_content)
                        print("Created")
                    time.sleep(2)
                    # display in browser
                    # ------------------
                    webbrowser.open_new_tab("report\\filt_report_index.html")
                    print('')
                    wait = input("Press ENTER to return")
            # filter by type "search"
            #----------------------
            elif filt == "2":
                cqn = '"search"'
                print(Fore.GREEN + 'Filtered Report Index')
                print(Fore.GREEN + '----------------------')
                print(Style.RESET_ALL)
                print('')     
                mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value\
                                        WHERE cqType_name = " + cqn)
                cursor = CONNECTION.cursor(buffered = True)
                cursor.execute(mysql_search_query)    
                mytable = from_db_cursor(cursor)
                mytable.align = "l"
                print(mytable)
                print('')
                # send report to browser
                # -----------------------
                printRep = input(Fore.YELLOW + 
                                'To send this report to the browser for printing'
                                ' or saving enter b or B, otherwise press '
                                'enter to return: ')
                print(Style.RESET_ALL)
                if printRep == "b" or printRep == "B":
                    # generate data for report
                    # ------------------------
                    mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value WHERE \
                                        cqType_name = " + cqn)
                    cursor = CONNECTION.cursor(buffered = True)
                    cursor.execute(mysql_search_query)
                    mytable = pd.read_sql("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value WHERE \
                                        cqType_name = " + cqn, CONNECTION)
                    pd.set_option('display.expand_frame_repr', False)
                    mytable2 = build_table(mytable,
                                         'grey_light',
                                         font_size = 'small',
                                         font_family = 'Open Sans, courier',
                                         text_align = 'left ')
                    # generate html content
                    # ---------------------
                    html_content = f"<html> \
                                    <head> <h2> ISBN22 Filtered Report Index \
                                    <br>Filter: {cqn} \
                                    </h2> \
                                    <h3> <script>\
                                    var timestamp = Date.now();\
                                    var d = new Date(timestamp);\
                                    document.write(d);\
                                    </script>\
                                    </h3>\
                                    </head> \
                                    <body> {mytable2} \
                                    </body> \
                                    </html>"
                    with open("report/filt_report_index.html", "w") as html_file:
                        html_file.write(html_content)
                        print("Created")
                    time.sleep(2)
                    # display in browser
                    # ------------------
                    webbrowser.open_new_tab("report\\filt_report_index.html")
                    print('')
                    wait = input("Press ENTER to return")
            # filter by type "list"
            #----------------------
            elif filt == "3":
                cqn = '"list"'
                print(Fore.GREEN + 'Filtered Report Index')
                print(Fore.GREEN + '----------------------')
                print(Style.RESET_ALL)
                print('')     
                mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value\
                                        WHERE cqType_name = " + cqn)
                cursor = CONNECTION.cursor(buffered = True)
                cursor.execute(mysql_search_query)    
                mytable = from_db_cursor(cursor)
                mytable.align = "l"
                print(mytable)
                print('')
                # send report to browser
                # -----------------------
                printRep = input(Fore.YELLOW + 
                                'To send this report to the browser for printing'
                                ' or saving enter b or B, otherwise press '
                                'enter to return: ')
                print(Style.RESET_ALL)
                if printRep == "b" or printRep == "B":
                    # generate data for report
                    # ------------------------
                    mysql_search_query = ("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value WHERE \
                                        cqType_name = " + cqn)
                    cursor = CONNECTION.cursor(buffered = True)
                    cursor.execute(mysql_search_query)
                    mytable = pd.read_sql("SELECT cq_name AS \"report name\", \
                                        cq_desc AS \"report description\", \
                                        cqType_name AS \"report type\", \
                                        cq_parameters AS \
                                        \"parameters(0=no, 1=yes)\",\
                                        cq_creator AS \"report orginator\", \
                                        cq_created AS \"origination date\" \
                                        FROM cQuery p INNER JOIN cqType c \
                                        ON p.cq_type = c.cqType_value WHERE \
                                        cqType_name = " + cqn, CONNECTION)
                    pd.set_option('display.expand_frame_repr', False)
                    mytable2 = build_table(mytable,
                                         'grey_light',
                                         font_size = 'small',
                                         font_family = 'Open Sans, courier',
                                         text_align = 'left ')
                    # generate html content
                    # ---------------------
                    html_content = f"<html> \
                                    <head> <h2> ISBN22 Filtered Report Index \
                                    <br>Filter: {cqn} \
                                    </h2> \
                                    <h3> <script>\
                                    var timestamp = Date.now();\
                                    var d = new Date(timestamp);\
                                    document.write(d);\
                                    </script>\
                                    </h3>\
                                    </head> \
                                    <body> {mytable2} \
                                    </body> \
                                    </html>"
                    with open("report/filt_report_index.html", "w") as html_file:
                        html_file.write(html_content)
                        print("Created")
                    time.sleep(2)
                    # display in browser
                    # ------------------
                    webbrowser.open_new_tab("report\\filt_report_index.html")
                    print('')
                    wait = input("Press ENTER to return")         
        # display report by name
        # -----------------------
        elif menuOption == '3':
            os.system('cls')
            now = datetime.datetime.now()
            print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
            print(("isbn-22 " + VERNO + " " + VERNA).rjust(80))
            print("-----------------------".rjust(80))
            print(Style.RESET_ALL)
            report = input("Please enter the name of the report you wish to "
                            "display: ")
            report2 = report.strip('View')
            cqdesc = ("SELECT cq_desc FROM cQuery WHERE cq_name = " + report2)
            cursor2 = CONNECTION.cursor(buffered = True)
            cursor2.execute(cqdesc)  
            cqresult = cursor2.fetchall()
            cq_filter = ("SELECT cq_filter FROM cQuery WHERE cq_name = " + report2)
            cursor3 = CONNECTION.cursor(buffered = True)
            cursor3.execute(cq_filter)  
            cqFilter = cursor3.fetchall()
            customQuery = ("SELECT * FROM " + report)
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(customQuery)    
            mytable = from_db_cursor(cursor)
            mytable.align = "l"
            print('')
            print(mytable)
            print('')
            printRep = input(Fore.YELLOW + 'To send this report to the '
                                'browser for printing or saving enter b  '
                                'or B, otherwise press enter to return: ')
            print(Style.RESET_ALL)
            # send report to browser
            # -----------------------
            if printRep == "b" or printRep == "B":
                mysql_search_query = (customQuery)
                cursor = CONNECTION.cursor(buffered = True)
                cursor.execute(mysql_search_query)
                mytable = pd.read_sql(customQuery, CONNECTION)
                pd.set_option('display.expand_frame_repr', False)
                mytable2 = build_table(mytable,
                                     'grey_light',
                                     font_size = 'small',
                                     font_family = 'Open Sans, courier',
                                     text_align = 'left ')
                # generate html content
                # ---------------------
                html_content = f"<html> \
                                <head> \
                                <h2>ISBN22 Report {report} - Descrition: \
                                {cqresult}<br>Filter: {cqFilter} \
                                </h2> \
                                <h3> <script>\
                                var timestamp = Date.now();\
                                var d = new Date(timestamp);\
                                document.write(d);\
                                </script>\
                                </h3>\
                                </head> \
                                <body> {mytable2} \
                                </body> \
                                </html>"
                with open("report/" + report + "_custom.html", "w") as html_file:
                    html_file.write(html_content)
                    print("Created")
                time.sleep(2)
                # display in browser
                # ------------------
                webbrowser.open_new_tab("report\\" + report + "_custom.html")
                print('')
        # create a new report (create a custom query)
        # --------------------------------------------
        if menuOption == '4':
            os.system('cls')
            print(Fore.GREEN + 'CREATE NEW REPORT QUERY')
            print(Fore.GREEN + '-------------------------')
            print(Style.RESET_ALL)
            print('')
            # define custom query
            # -------------------    
            customQuery = input("Please enter your query here,"
            " note, query must follow valid SQL syntax:   ")
            customQuery = customQuery.strip(';')
            # test SQL syntax
            # ----------------
            if testSQL(customQuery) == 1:
                print('')
                print(Fore.RED + "There is an error in you MySQL syntax")
                print('Please try again')
                print(Style.RESET_ALL)
            else:
                cursor = CONNECTION.cursor(buffered = True)
                cursor.execute(customQuery)    
                mytable = from_db_cursor(cursor)
                mytable.align = "l"
                print('')
                print(mytable)
                print('')
                printRep = input(Fore.YELLOW + 'To send this report to the '
                                    'browser for printing or saving enter b  '
                                    'or B, otherwise press enter to return: ')
                print(Style.RESET_ALL)
                # send report to browser
                # -----------------------
                if printRep == "b" or printRep == "B":
                    mysql_search_query = (customQuery)
                    cursor = CONNECTION.cursor(buffered = True)
                    cursor.execute(mysql_search_query)
                    mytable = pd.read_sql(customQuery, CONNECTION)
                    pd.set_option('display.expand_frame_repr', False)
                    mytable2 = build_table(mytable,
                                         'grey_light',
                                         font_size = 'small',
                                         font_family = 'Open Sans, courier',
                                         text_align = 'left ')
                    # generate html content
                    # ---------------------
                    html_content = f"<html> \
                                    <head> \
                                    <h2>ISBN22 Report (Prefactory) - Filter: \
                                    {customQuery} \
                                    </h2> \
                                    <h3> <script>\
                                    var timestamp = Date.now();\
                                    var d = new Date(timestamp);\
                                    document.write(d);\
                                    </script>\
                                    </h3>\
                                    </head> \
                                    <body> {mytable2} \
                                    </body> \
                                    </html>"
                    with open("report/prefactory_custom.html", "w") as html_file:
                        html_file.write(html_content)
                        print("Created")
                    time.sleep(2)
                    # display in browser
                    # ------------------
                    webbrowser.open_new_tab("report\\prefactory_custom.html")
                    print('')
                # option to save custom query to report index
                # --------------------------------------------
                print("The custom query was \n\n" + Fore.RED + customQuery + 
                        Style.RESET_ALL + "\n\nWould you like to save this"
                        " query?")
                saveQ = input("If you want to save this query, enter y or Y: ")
                if saveQ == 'y' or saveQ == 'Y':
                    saveQdesc = input("Please enter a brief description of the "
                                "query: ")
                    print('')
                    # define if parameters need to entered for query
                    # -----------------------------------------------
                    para = input("Does this query require user-defined "
                                "parameters?\n"
                                "Enter 'y' or 'Y' for yes: ")
                    if para == 'y' or para == 'Y':
                        cqPara = 1
                    else:
                        cqPara = 0
                    print('')
                    # display options for query types
                    # ================================       
                    # MySQL query to update VIEW if types are added
                    # ==============================================
                    # CREATE OR REPLACE VIEW cqTypeView AS SELECT cqType_value 
                    # AS "value", cqType_name AS "name", cqType_desc 
                    # AS "description" FROM cqType;
                    goAgain3 = 1
                    while goAgain3 == 1:
                        print("Available types are:")
                        print("--------------------")
                        print('')
                        mysql_search_query = ("SELECT cqType_value, cqType_name,\
                                            cqType_desc FROM cqType")
                        cursor = CONNECTION.cursor(buffered = True)
                        cursor.execute(mysql_search_query)    
                        mytable = from_db_cursor(cursor)
                        mytable.align = "l"
                        print(mytable)
                        print('')
                        print('')                       
                        saveQtype = input("Please enter a value for the query "
                                        "type: ") 
                        # require user selection to match one of the available 
                        # types
                        # ======================================================
                        # action to follow is selection matches available type
                        # ======================================================     
                        if saveQtype == '1' or saveQtype == '2' or \
                            saveQtype == '3':
                            #    goAgain2 = 0           
                            today = date.today()
                            mysql_cqid_query = ("SELECT cq_id FROM cQuery ORDER \
                                                 BY cq_id\
                                                 DESC LIMIT 1")
                            cursor = CONNECTION.cursor(buffered = True)
                            cursor.execute(mysql_cqid_query)
                            result = cursor.fetchall()
                            for row in result:
                                cqid = (row[0]) 
                                cqid += 1
                            saveQname = str(cqid)
                            nq = ("SELECT * FROM " + saveQname + "View")
                            createView = ("CREATE OR REPLACE VIEW " + \
                                        saveQname + "View AS " + customQuery)
                            cursor = CONNECTION.cursor()
                            cursor.execute(createView)
                            CONNECTION.commit()
                            data = (saveQname, saveQdesc, saveQtype, customQuery, \
                                    nq, USER, today, cqPara)
                            cq_insert_query = ("INSERT INTO cQuery (cq_name, \
                                            cq_desc, cq_type,cq_filter, cq_query, \
                                            cq_creator, cq_created, cq_parameters)"
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)")     
                            cursor = CONNECTION.cursor()
                            cursor.execute(cq_insert_query, data)
                            CONNECTION.commit()
                            print('')
                            print(Fore.YELLOW + "cQuery "+ saveQname +  \
                                "View created")
                            print(Style.RESET_ALL)
                            goAgain3 = 0
                        # action to follow if selection does not match available 
                        # types (notify user and return them to type selection)
                        # ======================================================    
                        else:
                            print(Fore.RED + 'Please make a valid selection')
                            cancelSelc = input("Press any key to continue or 0 to"
                                                " CANCEL: ")
                            print(Style.RESET_ALL)
                            if cancelSelc == '0':
                                goAgain3 = 0 


                


        elif menuOption == '0':    
            goAgain = 0                 


def searchService():
    """
    ============================================================================
    Function:       searchService()
    Purpose:        defines the online service to use for isbn data
    Parameter(s):   -None- 
    Return:         -None- (value written to variable)
    ============================================================================
    """
    global SERVICE

    print('Search Service Options:')
    print('------------------------')
    print('1\tGoogle Books (goob)')
    print('2\tWikipedia (wiki)')
    print('3\tOpenLibrary (openl)') 
    print('')
    print('')
    service = input("Select Service: ")
    if service == '1':
        SERVICE = "goob"
    elif service == '2':
        SERVICE = "wiki"
    elif service == '3':
        SERVICE = "openl"
    print(Fore.YELLOW + "Service Set To: " + SERVICE)
    print(Style.RESET_ALL)

def testSQL(tquery):
    """
    ============================================================================
    Function:       testSQL(tquery)
    Purpose:        tests that a MySQL query functions and is written in the 
                    correct syntax
    Parameter(s):   MySQL query (tyquery)
    Return:         1 if tquery contains a MySQL error, nothing if not
    ============================================================================
    """
    test_query = (tquery)
    try:
        cursor = CONNECTION.cursor(buffered = True)
        cursor.execute(test_query)
    except mysql.connector.errors.ProgrammingError as err:
        if err:
            pass
            return (1)

# ==============================================================================
#  main entry point for program
#  =============================================================================    

def main():
    """
    ============================================================================
    Function:       main()
    Purpose:        entry point to program
    Parameter(s):   -None-
    Return:         -None-
    ============================================================================
    """
    global CONNECTION
    #preProcess()
    #createLists()
    #getInfo()
    #getGenre()
    #exportLists()
    menu()
    print("Closing Database Connection . . .")
    CONNECTION.close()
    print("bye . . .")

if __name__ == "__main__":
    main()

