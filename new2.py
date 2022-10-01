import sys
#from isbntools.app import *
import pandas as pd
from isbnlib import meta
from isbnlib.registry import bibformatters
import openpyxl
import mysql.connector
from mysql.connector import Error
import functools
import xlsxwriter
import pandas.io.sql as sql


workbookName = "inventory.xlsx"
dataframeName = "dataframe.xlsx"
dbIsbnxls = "dbxls.xlsx"
# load excel with its path
wrkbk = openpyxl.load_workbook(workbookName)
#dframe = openpyxl.load_workbook(dataframeName)
  
sh = wrkbk.active



# initialize lists
bad_list = []
good_list = []
genre_list = []
isbn_list = []
dbIsbn_list = []

try:
    connection = mysql.connector.connect(user='jfsharron', password='marie151414',
    host='192.168.2.107', database='isbn22')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = connection.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

def preProcess():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """
    
    print("Checking for duplicates in source file . . .")

    #for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row, max_col=2):
    #    for cell in row:
    #        isbn = str(cell.value)
    #        isbn_list.append(isbn)

    # remove duplicates from isbn spreadsheet, save in dataframe spreadsheet,
    # import into isbn_list
    # ==========================================================================
    data = pd.read_excel(workbookName, usecols = ['isbn'])
    data_first_record = data.drop_duplicates(keep="first")

    writer = pd.ExcelWriter(dataframeName, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    data_first_record.to_excel(writer, sheet_name='Sheet1')

    # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # Add some cell formats.
    format1 = workbook.add_format({'num_format': '###0'})

    worksheet.set_column(1, 1, 18, format1)

    writer.save()

    dframe = openpyxl.load_workbook(dataframeName)
    sh = dframe.active

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            isbn = str(cell.value)
            isbn_list.append(isbn)
    
     # create dbIsbn_list from database
    # ==========================================================================
    print("Checking for duplicates in database file . . .")
    query = "SELECT isbn FROM isbn"
    df = sql.read_sql('SELECT isbn FROM isbn', connection)
    print(df)
    #dbIsbn_list.append(df)
    df.to_excel(dbIsbnxls)

    dbexcel = openpyxl.load_workbook(dbIsbnxls)
    sh = dbexcel.active

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            isbn = str(cell.value)
            dbIsbn_list.append(isbn)
    
    
    
    
    
    
    #cursor.execute(query)
    #rows = cursor.fetchall()
    #connection.commit()
    ##print(rows)
    #dbIsbn_list.append(rows)
    #
    # compare lists and create intersection list
    # ==========================================================================
    a = (isbn_list)
    #print("TYPE a")
    #print(type(a))
    #print(a)
    b = (dbIsbn_list)
    #print("TYPE b")
    #print(type(b))
    #print(b)

    # ===========================================================================
    print("LISTS BEFORE INTERSECTION")
    print(isbn_list)
    print()
    print(dbIsbn_list)
    # ===========================================================================

    intersection = set(a).intersection(b)
    print("Intersection")
    print(list(intersection))

    

    # remove intersection list values from isbn_list
    # ==========================================================================
    print("BEFORE")
    print(isbn_list)
    
    for value in intersection:
        if value in isbn_list:
            #print("exists")
            isbn_list.remove(value)
    print("AFTER")
    print(isbn_list)

#===============================================================================    







def get_difference(list_a, list_b):
    intersection = set(list_a.intersection(list_b))
    #non_match_a  = set(list_a)-set(list_b)
    #non_match_b  = set(list_b)-set(list_a)
    #non_match = list(non_match_a) + list(non_match_b)
    return intersection


    
    






    # ==========================================================================

# Input Variables
def createLists():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """

    # iterate through excel and display data
    for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row, max_col=1):
        for cell in row:
            isbn = str(cell.value)

            SERVICE = "openl"

            bibtex = bibformatters["bibtex"]
            
            meta_dict = meta(isbn, service='default')

            if meta_dict.get('Authors') is None:
                bad_list.append(isbn)
            else:
                good_list.append(isbn)

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            genre = str(cell.value)
            genre_list.append(genre)

    #print("bad list ", bad_list)
    #print("good list ", good_list)
    #print("genre list ", genre_list)

    # ==========================================================================








def main():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """
    #dbConnect()
    preProcess()
    #createLists()
    #getInfo()
    #exportBad()
    #exportDb()
    #print(good_list)
    #get_difference(isbn_list, dbIsbn_list)
#
    #intersection = get_difference(isbn_list, dbIsbn_list)
    #print("Non-match elements: ", intersection)

    print("Closing Database Connection . . .")
    cursor.close()
    connection.close()
    print("bye . . .")



if __name__ == "__main__":
    main()
