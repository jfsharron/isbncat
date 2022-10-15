"""
================================================================================
 Program:           newMain.py
 Software Engineer: Jonas SHarron
 Date:              01-October-2022

 Purpose:   This program will process isbn's stored in a file and export them to 
            a MySQL database.  The program will also check for isbn's that are
            not represented in the search service and export them to list for
            evaluation by the user.  A manuel entry and editing method is also 
            provided for user interaction.
================================================================================
"""

import sys
import pandas as pd
from isbnlib import meta
from isbnlib.registry import bibformatters
import openpyxl
import mysql.connector
from mysql.connector import Error
import functools
import xlsxwriter
import pandas.io.sql as sql
import xlrd
import csv
#import dataconverters.commas as commas
#import urllib3 as urllib2



# define external files
# ======================
workbookName = "inventory.xlsx"
dataframeName = "dataframe.xlsx"
dbIsbnxls = "dbxls.xlsx"

# initialize lists
# =================
bad_list = []
good_list = []
genre_list = []
isbn_list = []
dbIsbn_list = []
dup_list = []
#isbn_dict = {"isbn", "genre"}

# establish database connection
# =============================
try:
    connection = mysql.connector.connect(user='jfsharron', password='marie151414',
    host='192.168.2.107', database='isbn22')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = connection.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

# ==============================================================================
# user functions
# ==============================================================================

def createLists():
    """
    ============================================================================
    Function:       createLists()
    Purpose:        filters input data into two lists (good_list and bad_list)
                    depending on the availability of data in the search service
                    (filtered by return data from ('Authors') field)
    Parameter(s):   -None- (processes data in external file)
    Return:         -None- (propagates data in good_list and bad_list)
    ============================================================================
    """
     # iterate through data file and for existence in search service
    print("Checking search service for data file information")
    
    for value in isbn_dict.values():
        isbn = value

    #for k, v in isbn_dict.items():
    #    if k == "isbn":
    #        print(k)
#

    #for i in isbn_dict:
    #    isbn = isbn_dict.isbn()

        SERVICE = "openl"
        bibtex = bibformatters["bibtex"]
        
        # if data is not available in search service, add isbn to bad_list
        # otherwise add isbn to good_list
        # ==================================================================
        meta_dict = meta(isbn, service='default')
        if meta_dict.get('Authors') is None:
            bad_list.append(isbn)
        else:
            good_list.append(isbn)  

    # data file check completion message
    print("Check for datafile information completed, data with available "
        "information exported to good_list, data without information exported"
        " to bad_list")

def getInfo():    
    """
    ============================================================================
    Function:       getInfo()
    Purpose:        retrieves information (author, title, isbn, year, publisher)
                    from search service and propagates MySQL database with hew
                    record.
    Parameter(s):   -None- (processes data in good_list)
    Return:         -None- (exports data to MySQL database)
    ============================================================================
    """
    # iterate through good_list and retrieve data from search service
    print("Connecting to search service . . .")
    print("Retrieving information for good_list . . .")
    for i in good_list:

            isbn = i

            SERVICE = "openl"

            bibtex = bibformatters["bibtex"]
            
            meta_dict = meta(isbn, service='default')

            aut = str(meta_dict['Authors'])
            aut = aut.replace("[","")
            aut = aut.replace("]","")
            author = aut.replace("'","")
            title = meta_dict['Title']
            isbn = meta_dict['ISBN-13']
            year = meta_dict['Year']
            publisher = meta_dict['Publisher']

            # define data to retrieve
            data = (isbn, year, publisher, author, title)

            # SQL query to insert data into to db
            mySql_insert_query = (
            "INSERT INTO isbn (isbn, year, publisher, author, title)"
            "VALUES (%s, %s, %s, %s, %s)"
            )
        
            cursor = connection.cursor()
            cursor.execute(mySql_insert_query, data)
            connection.commit()

    # completion message
    print("Available information added to SQL database")            
            

# ==============================================================================

def exportBad():
    """
    ============================================================================
    Function:       exportBad()
    Purpose:        exports bad list (isbn's not found in search service) to 
                    external file for evaluation
    Parameter(s):   -None- (processes data in bad_list)
    Return:         -None- (generates external file)    
    ============================================================================
    """
    with open(r'bad_list.txt', 'w') as fp:
        for i in bad_list:
            fp.write("%s\n" % i)
    
    fp.close()

def preProcess():
    """
    ============================================================================
    Function:       preProcess()
    Purpose:        looks for and removes duplicate isbn's from import file and 
                    pre-existing db 
    Parameter(s):   -None- (reads data from input file and db)
    Return:         -None- (writes data (duplicates removed) to isbn_list)
    ============================================================================
    """
    print("Checking for duplicates in source file . . .")


    # remove duplicates from isbn spreadsheet, save in dataframe spreadsheet
    data = pd.read_excel(workbookName, usecols = ['isbn', 'genre'])
    data_first_record = data.drop_duplicates(keep="first")

    writer = pd.ExcelWriter(dataframeName, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    data_first_record.to_excel(writer, sheet_name='Sheet1')

    # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # Add numeric cell formats.
    format1 = workbook.add_format({'num_format': '###0'})

    worksheet.set_column(1, 1, 18, format1)

    writer.save()

    read_file = pd.read_excel(dataframeName)

    read_file.to_csv("isbn.csv", index = None, header = True)

    csvdf = pd.read_csv("isbn.csv")

    with open("isbn.csv", "r") as source:
        reader = csv.reader(source)

        with open("output.csv", "w") as result:
            writer = csv.writer(result)
            for r in reader:

                format1 = workbook.add_format({'num_format': '###0'})   
                worksheet.set_column(2, 2, 18, format1) 
                writer.writerow((r[1], r[2]))

    output = pd.read_csv("output.csv")

    #print(csvdf)

    #csvdf.drop("Unnamed: 0", inplace = True, axis = 1)

    #read_file.to_csv("isbn2.csv", index = False, header = True)

    #isbn_dict = pd.DataFrame(data, columns = ['isbn', 'genre'])

    isbn_dict = output.to_dict()

    print(isbn_dict)

    #filename = 'isbn.csv'
    #with open(filename) as f:
    #    f_writer = csv.writer
#
#
#
    #    #records, metadata = commas.parse(f)
    #    for row in f:
    #        print("this is row in dictionary:"+row)



    #wb = xlrd.open_workbook(dataframeName)
    #sh = writer.sheets['Sheet1']
    #isbn = isbn_dict((sh.col_values(1, 0, 138), (sh.col_values(2, 0, 138))))


    #for i in range(64):
    #    cell_value_class = sh.cell(i,2).value
    #    cell_value_id = sh.cell(i,0).value
    #    d[cell_value_class] = cell_value_id
 
    #xls = dataframeName
    #data = xls.parse(xls.sheet_names[0])
    #print(data.to_dict())


    #df = pd.read_excel(dataframeName)

    ###for row in df:
    ###    isbn_dict = df.to_dict('isbn', 'genre')
    ###print(isbn_dict)

    #dframe = openpyxl.load_workbook(dataframeName)
    #sh = dframe.active
#
    ## send dataframe data to isbn_dict
    #dframe.to_dict(isbn_dict)


    #for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
    #    for cell in row:
    #        isbn = str(cell.value)
    #        isbn_list.append(isbn)
    
    # create dbIsbn_list from database
    print("Checking for duplicates in database file . . .")
    query = "SELECT isbn FROM isbn"
    df = sql.read_sql('SELECT isbn FROM isbn', connection)
    df.to_excel(dbIsbnxls)

    dbexcel = openpyxl.load_workbook(dbIsbnxls)
    sh = dbexcel.active

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            isbn = str(cell.value)
            dbIsbn_list.append(isbn)
    
    # compare isbn_list and dbIsbn_list and create intersection (duplicates) list
    a = (isbn_dict)
    b = (dbIsbn_list)


    intersection = set(a).intersection(b)
    

    # remove intersection (duplicates) list values from isbn_list
    for value in intersection:
        if value in isbn_list:
            dup_list.append(value)
            isbn_list.remove(value)

    # completion of duplicate check message
    print("Duplicate check completed, duplicates removed and exported to dup_list")

# ==============================================================================
#  main entry point for program
#  =============================================================================    

def main():
    """
    ============================================================================
    Function:       main()
    Purpose:        entry point to program
    Parameter(s):   -None-
    Return:         -None-
    ============================================================================
    """
    preProcess()
    createLists()
    #getInfo()
    #exportBad()
    #print("Closing Database Connection . . .")
    connection.close()
    print("bye . . .")

if __name__ == "__main__":
    main()

