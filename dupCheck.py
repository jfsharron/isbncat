import sys
#from isbntools.app import *
import pandas as pd
from isbnlib import meta
from isbnlib.registry import bibformatters
import openpyxl
import mysql.connector
from mysql.connector import Error
import functools


workbookName = "inventory.xlsx"
# load excel with its path
wrkbk = openpyxl.load_workbook(workbookName)
  
sh = wrkbk.active



# initialize lists
bad_list = []
good_list = []
genre_list = []
isbn_list = []
dbIsbn_list = []

try:
    connection = mysql.connector.connect(user='jfsharron', password='marie151414',
    host='192.168.2.107', database='isbn22')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = connection.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

def preProcess():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """
    
    print("Checking for duplicates in source file . . .")
    #isbn_list.append(isbn)
    #print("isbn list")
    #print(isbn_list)
    ##res = [*set(isbn1)]
    ##for key, value in dict.iteritems():
    ##    temp = [key,value]
    ##    dictlist.append(temp)
#
    #
    #
    #
    #
    #
    #
    #
    #
    #
    #
    #
    #
    data = pd.read_excel(workbookName, usecols = ['isbn'])
    data_first_record = data.drop_duplicates(keep="first")
    print(data_first_record)
    isbn_list = data_first_record.values.tolist()
    ##print()
    ##print()
    ##print()
    ##print()
    ##print(isbn_list)
    ##print()
    ##print()
    ##print()
    ##print()
    ##print(dbIsbn_list)
#
    print("Checking for duplicates in database file . . .")
    query = "SELECT isbn FROM isbn"
    cursor.execute(query)
    rows = cursor.fetchall()
    connection.commit()
    print(rows)
    dbIsbn_list.append(rows)

    print("db list2")
    print(dbIsbn_list)

    #for record in records:
    #    print(record)
    #dbIsbn_list = records.values.tolist()
    #print(dbIsbn_list)
    #x = set(isbn_list) & set(dbIsbn_list)
    #set(isbn_list).intersection(set(dbIsbn_list))
    #print(x)
    #a = isbn_list
    #b = dbIsbn_list 
    print("TYPE ISBN")
    print(type(isbn_list))
    print("TYPE DB")
    print(type(dbIsbn_list))

    a = tuple(isbn_list)
    print("TYPE a")
    print(type(a))
    print(a)
    b = tuple(dbIsbn_list)
    print("TYPE b")
    print(type(b))
    print(b)

    

    #a = (isbn_list)
    ##a = [1,3,4]
    #b = (dbIsbn_list)  
    #print("isbn")
    #print(a)
    #print()
    #print("dbisbn")
    #print(b)
    #
    intersection = set(a).intersection(b)
    print(list(intersection))







def get_difference(list_a, list_b):
    intersection = set(list_a.intersection(list_b))
    #non_match_a  = set(list_a)-set(list_b)
    #non_match_b  = set(list_b)-set(list_a)
    #non_match = list(non_match_a) + list(non_match_b)
    return intersection


    
    






    # ==========================================================================

# Input Variables
def createLists():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """

    # iterate through excel and display data
    for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row, max_col=1):
        for cell in row:
            isbn = str(cell.value)

            SERVICE = "openl"

            bibtex = bibformatters["bibtex"]
            
            meta_dict = meta(isbn, service='default')

            if meta_dict.get('Authors') is None:
                bad_list.append(isbn)
            else:
                good_list.append(isbn)

    for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
        for cell in row:
            genre = str(cell.value)
            genre_list.append(genre)

    #print("bad list ", bad_list)
    #print("good list ", good_list)
    #print("genre list ", genre_list)

    # ==========================================================================








def main():
    """
    ============================================================================
    Function:
    Purpose:
    Parameter(s):
    Return:

    creates good_list and bad_list from imported .xlsx file
    ============================================================================
    """
    #dbConnect()
    preProcess()
    #createLists()
    #getInfo()
    #exportBad()
    #exportDb()
    #print(good_list)
    #get_difference(isbn_list, dbIsbn_list)
#
    #intersection = get_difference(isbn_list, dbIsbn_list)
    #print("Non-match elements: ", intersection)

    print("Closing Database Connection . . .")
    cursor.close()
    connection.close()
    print("bye . . .")

if __name__ == "__main__":
    main()
